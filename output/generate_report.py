"""
Phase 3: Generate portfolio report from extracted screenshot data.
Outputs portfolio_detail.csv, portfolio_summary.csv, and portfolio_report.xlsx.
"""
import csv
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

OUTPUT_DIR = Path(__file__).parent
TIMESTAMP = datetime.now().strftime("%Y-%m-%d %H:%M")

# ── Extracted data from screenshots ──────────────────────────────

DETAIL_ROWS = [
    # wallet_address, chain, category, protocol, token_symbol, token_name, amount, usd_value

    # === 0x040043 ($9,310) ===
    # Wallet tokens
    ("0x0400435d6d20099a19FA138F74Ef46810478A8db", "Base", "wallet", "-", "ETH", "Ethereum", 0.2745, 530.64),
    ("0x0400435d6d20099a19FA138F74Ef46810478A8db", "Base", "wallet", "-", "USDC", "USD Coin", 405.5009, 405.58),
    ("0x0400435d6d20099a19FA138F74Ef46810478A8db", "Ethereum", "wallet", "-", "ETH", "Ethereum", 0.1208, 233.51),
    ("0x0400435d6d20099a19FA138F74Ef46810478A8db", "Arbitrum", "wallet", "-", "ETH", "Ethereum", 0.0806, 156.04),
    ("0x0400435d6d20099a19FA138F74Ef46810478A8db", "Avalanche", "wallet", "-", "AVAX", "Avalanche", 9.2568, 80.26),
    ("0x0400435d6d20099a19FA138F74Ef46810478A8db", "Sonic", "wallet", "-", "S", "Sonic", 1527.7688, 61.61),
    ("0x0400435d6d20099a19FA138F74Ef46810478A8db", "Ethereum", "wallet", "-", "USDC", "USD Coin", 20.0744, 20.08),
    ("0x0400435d6d20099a19FA138F74Ef46810478A8db", "Arbitrum", "wallet", "-", "USDC", "USD Coin", 13.9376, 13.94),
    ("0x0400435d6d20099a19FA138F74Ef46810478A8db", "Avalanche", "wallet", "-", "USDC", "USD Coin", 9.9975, 10.00),
    ("0x0400435d6d20099a19FA138F74Ef46810478A8db", "BNB Chain", "wallet", "-", "BTCB", "Bitcoin BEP2", 0.00008668, 5.73),
    ("0x0400435d6d20099a19FA138F74Ef46810478A8db", "Polygon", "wallet", "-", "ETH", "Ethereum", 0.0026, 5.11),
    ("0x0400435d6d20099a19FA138F74Ef46810478A8db", "BNB Chain", "wallet", "-", "BUSD", "Binance USD", 2.8036, 2.81),
    ("0x0400435d6d20099a19FA138F74Ef46810478A8db", "Polygon", "wallet", "-", "POL", "Polygon", 21.9603, 2.55),
    ("0x0400435d6d20099a19FA138F74Ef46810478A8db", "Base", "wallet", "-", "USDbC", "USD Base Coin", 2.5011, 2.50),
    ("0x0400435d6d20099a19FA138F74Ef46810478A8db", "Base", "wallet", "-", "cbBTC", "Coinbase BTC", 0.00003672, 2.43),
    ("0x0400435d6d20099a19FA138F74Ef46810478A8db", "Ethereum", "wallet", "-", "WETH", "Wrapped ETH", 0.0012, 2.30),
    ("0x0400435d6d20099a19FA138F74Ef46810478A8db", "BNB Chain", "wallet", "-", "BNB", "BNB", 0.0028, 1.69),
    ("0x0400435d6d20099a19FA138F74Ef46810478A8db", "Arbitrum", "wallet", "-", "ETH", "Ethereum", 0.0009, 1.68),
    # DeFi
    ("0x0400435d6d20099a19FA138F74Ef46810478A8db", "Base", "defi", "Avantis", "USDC", "USD Coin (Yield)", 7764.4432, 7766.00),

    # === 0x1E2b54 ($2,451) ===
    # Wallet tokens
    ("0x1E2b54e08767A407C7EBef8FFF4142cE0D92ce16", "HyperEVM", "wallet", "-", "HYPE", "Hyperliquid", 0.3501, 9.68),
    ("0x1E2b54e08767A407C7EBef8FFF4142cE0D92ce16", "HyperEVM", "wallet", "-", "ETH", "Ethereum", 0.0012, 2.28),
    ("0x1E2b54e08767A407C7EBef8FFF4142cE0D92ce16", "HyperEVM", "wallet", "-", "ETH", "Ethereum", 0.0010, 1.92),
    # DeFi - Hyperliquid
    ("0x1E2b54e08767A407C7EBef8FFF4142cE0D92ce16", "Hyperliquid", "defi", "Hyperliquid", "USDC", "USD Coin (Spot)", 1861.2541, 1861.25),
    ("0x1E2b54e08767A407C7EBef8FFF4142cE0D92ce16", "Hyperliquid", "defi", "Hyperliquid", "HYPE", "Hyperliquid (Spot)", 20.8184, 575.37),

    # === 0xFE10ce ($991) ===
    # DeFi - Hyperliquid
    ("0xFE10ce58Be9Aa8cB8A3F88603E2D55e4178D19E9", "Hyperliquid", "defi", "Hyperliquid", "USDC", "USD Coin (Perps)", 990.1474, 990.15),

    # === 0xb38B29 ($4,215) ===
    # Wallet tokens
    ("0xb38B296Dbd09f881429D414b09A3294FFea2ad8c", "HyperEVM", "wallet", "-", "HYPE", "Hyperliquid", 4.8070, 132.84),
    ("0xb38B296Dbd09f881429D414b09A3294FFea2ad8c", "Arbitrum", "wallet", "-", "ETH", "Ethereum", 0.0176, 34.05),
    ("0xb38B296Dbd09f881429D414b09A3294FFea2ad8c", "HyperEVM", "wallet", "-", "ETH", "Ethereum", 0.0036, 7.01),
    # DeFi - Hyperliquid
    ("0xb38B296Dbd09f881429D414b09A3294FFea2ad8c", "Hyperliquid", "defi", "Hyperliquid", "USDC", "USD Coin (Spot)", 946.1582, 946.16),
    ("0xb38B296Dbd09f881429D414b09A3294FFea2ad8c", "Hyperliquid", "defi", "Hyperliquid", "FEUSD", "FEUSD (Spot)", 1972.9823, 1972.98),
    ("0xb38B296Dbd09f881429D414b09A3294FFea2ad8c", "Hyperliquid", "defi", "Hyperliquid", "FEUSD", "FEUSD (Spot Order)", 1119.0000, 1113.41),
    ("0xb38B296Dbd09f881429D414b09A3294FFea2ad8c", "Hyperliquid", "defi", "Hyperliquid", "USDH", "USDH (Spot)", 11.1674, 11.17),

    # === 0x76fD9c ($383) ===
    # Wallet tokens
    ("0x76fD9c9e75D91070947ad04A098C091e3f098c69", "Arbitrum", "wallet", "-", "ETH", "Ethereum", 0.0040, 7.70),
    ("0x76fD9c9e75D91070947ad04A098C091e3f098c69", "Ethereum", "wallet", "-", "ETH", "Ethereum", 0.0010, 1.93),
    # DeFi
    ("0x76fD9c9e75D91070947ad04A098C091e3f098c69", "Arbitrum", "defi", "LOCKON", "MULTI", "LMA Pool (ARB+AAVE+USDC+WETH+PENDLE)", 1, 194.37),
    ("0x76fD9c9e75D91070947ad04A098C091e3f098c69", "Avalanche", "defi", "Stable Jack", "xAVAX", "xAVAX (Yield)", 20.5519, 177.98),

    # === 7UPXsMLQ (Solana) $144.58 ===
    ("7UPXsMLQ8p2WeReYuQHN3wKpFMzrNQMkZ1BMrFcwbrhp", "Solana", "wallet", "-", "USDS", "USDS", 85.111, 85.11),
    ("7UPXsMLQ8p2WeReYuQHN3wKpFMzrNQMkZ1BMrFcwbrhp", "Solana", "wallet", "-", "SOL", "Solana", 0.235, 19.42),
    ("7UPXsMLQ8p2WeReYuQHN3wKpFMzrNQMkZ1BMrFcwbrhp", "Solana", "wallet", "-", "USDC", "USD Coin", 18.509, 18.51),
    ("7UPXsMLQ8p2WeReYuQHN3wKpFMzrNQMkZ1BMrFcwbrhp", "Solana", "wallet", "-", "TRUMP", "Official Trump", 4.212, 14.66),
    ("7UPXsMLQ8p2WeReYuQHN3wKpFMzrNQMkZ1BMrFcwbrhp", "Solana", "wallet", "-", "MELANIA", "Melania Meme", 60.0054, 6.81),
]

HEADERS = ["wallet_address", "chain", "category", "protocol", "token_symbol", "token_name", "amount", "usd_value", "timestamp"]


def spam_filter(row):
    """Filter out spam/scam tokens and dust."""
    symbol = row[4].lower()
    name = row[5].lower()
    usd = row[7]
    spam_patterns = ["claim", "airdrop", "visit", "redeem", "t.me", "t.ly", ".com", ".xyz", ".io"]
    for pat in spam_patterns:
        if pat in symbol or pat in name:
            return False
    if usd < 1.0:
        return False
    return True


def write_detail_csv():
    filtered = [r for r in DETAIL_ROWS if spam_filter(r)]
    path = OUTPUT_DIR / "portfolio_detail.csv"
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        for r in filtered:
            w.writerow(list(r) + [TIMESTAMP])
    print(f"Detail CSV: {path} ({len(filtered)} rows)")
    return filtered


def write_summary_csv(filtered):
    # Aggregate by token symbol
    agg = {}
    for r in filtered:
        sym = r[4]
        name = r[5].split(" (")[0]  # Remove annotation like "(Spot)", "(Yield)"
        if sym not in agg:
            agg[sym] = {"name": name, "amount": 0, "usd": 0}
        agg[sym]["amount"] += r[6]
        agg[sym]["usd"] += r[7]

    sorted_agg = sorted(agg.items(), key=lambda x: x[1]["usd"], reverse=True)

    path = OUTPUT_DIR / "portfolio_summary.csv"
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["token_symbol", "token_name", "total_amount", "total_usd_value"])
        for sym, data in sorted_agg:
            w.writerow([sym, data["name"], round(data["amount"], 6), round(data["usd"], 2)])
    print(f"Summary CSV: {path} ({len(sorted_agg)} tokens)")
    return sorted_agg


def write_xlsx(filtered, sorted_agg):
    if not HAS_OPENPYXL:
        print("openpyxl not installed, skipping XLSX")
        return

    wb = Workbook()

    # ── Summary Sheet ──
    ws = wb.active
    ws.title = "Summary"

    title_font = Font(size=16, bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    money_fmt = '#,##0.00'

    total_usd = sum(r[7] for r in filtered)

    ws["A1"] = "Portfolio Report"
    ws["A1"].font = title_font
    ws["A2"] = f"Generated: {TIMESTAMP}"
    ws["A4"] = "Total Portfolio Value (USD)"
    ws["A4"].font = Font(bold=True, size=14)
    ws["B4"] = total_usd
    ws["B4"].font = Font(bold=True, size=14, color="2E7D32")
    ws["B4"].number_format = '$#,##0.00'

    # Wallet breakdown
    row = 6
    ws.cell(row=row, column=1, value="Wallet Breakdown").font = Font(bold=True, size=12)
    row += 1
    for col, h in enumerate(["Wallet", "USD Value"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill

    wallet_totals = {}
    for r in filtered:
        addr = r[0]
        short = addr[:8] + "..."
        wallet_totals[short] = wallet_totals.get(short, 0) + r[7]

    row += 1
    for addr, val in sorted(wallet_totals.items(), key=lambda x: x[1], reverse=True):
        ws.cell(row=row, column=1, value=addr)
        c = ws.cell(row=row, column=2, value=val)
        c.number_format = money_fmt
        row += 1

    # Chain breakdown
    row += 1
    ws.cell(row=row, column=1, value="Chain Breakdown").font = Font(bold=True, size=12)
    row += 1
    for col, h in enumerate(["Chain", "USD Value", "Percentage"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill

    chain_totals = {}
    for r in filtered:
        chain = r[1]
        chain_totals[chain] = chain_totals.get(chain, 0) + r[7]

    row += 1
    for chain, val in sorted(chain_totals.items(), key=lambda x: x[1], reverse=True):
        ws.cell(row=row, column=1, value=chain)
        c = ws.cell(row=row, column=2, value=val)
        c.number_format = money_fmt
        pct = val / total_usd * 100 if total_usd > 0 else 0
        ws.cell(row=row, column=3, value=f"{pct:.1f}%")
        row += 1

    # Top 10
    row += 1
    ws.cell(row=row, column=1, value="Top 10 Holdings").font = Font(bold=True, size=12)
    row += 1
    for col, h in enumerate(["Token", "Name", "USD Value"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill

    row += 1
    for sym, data in sorted_agg[:10]:
        ws.cell(row=row, column=1, value=sym)
        ws.cell(row=row, column=2, value=data["name"])
        c = ws.cell(row=row, column=3, value=round(data["usd"], 2))
        c.number_format = money_fmt
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 15

    # ── Detail Sheet ──
    ws2 = wb.create_sheet("Detail")
    for col, h in enumerate(HEADERS, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill

    for i, r in enumerate(filtered, 2):
        for j, val in enumerate(list(r) + [TIMESTAMP], 1):
            c = ws2.cell(row=i, column=j, value=val)
            if j == 8:  # usd_value
                c.number_format = money_fmt

    for col in range(1, len(HEADERS) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 18
    ws2.column_dimensions["A"].width = 45

    # ── Token Summary Sheet ──
    ws3 = wb.create_sheet("Token Summary")
    tok_headers = ["Token", "Name", "Total Amount", "Total USD Value"]
    for col, h in enumerate(tok_headers, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill

    for i, (sym, data) in enumerate(sorted_agg, 2):
        ws3.cell(row=i, column=1, value=sym)
        ws3.cell(row=i, column=2, value=data["name"])
        ws3.cell(row=i, column=3, value=round(data["amount"], 6))
        c = ws3.cell(row=i, column=4, value=round(data["usd"], 2))
        c.number_format = money_fmt

    for col in range(1, 5):
        ws3.column_dimensions[get_column_letter(col)].width = 20

    path = OUTPUT_DIR / "portfolio_report.xlsx"
    wb.save(path)
    print(f"XLSX: {path}")


def main():
    print(f"=== Portfolio Report Generation ({TIMESTAMP}) ===\n")
    filtered = write_detail_csv()
    sorted_agg = write_summary_csv(filtered)
    write_xlsx(filtered, sorted_agg)

    # Print summary
    total = sum(r[7] for r in filtered)
    print(f"\n{'='*50}")
    print(f"  Total Portfolio Value: ${total:,.2f}")
    print(f"{'='*50}\n")

    # Per-wallet
    wallet_totals = {}
    for r in filtered:
        addr = r[0]
        short = addr[:8] + "..."
        wallet_totals[short] = wallet_totals.get(short, 0) + r[7]
    print("Wallet Breakdown:")
    for addr, val in sorted(wallet_totals.items(), key=lambda x: x[1], reverse=True):
        print(f"  {addr}  ${val:>10,.2f}")

    # Per-chain
    chain_totals = {}
    for r in filtered:
        chain_totals[r[1]] = chain_totals.get(r[1], 0) + r[7]
    print("\nChain Breakdown:")
    for chain, val in sorted(chain_totals.items(), key=lambda x: x[1], reverse=True):
        pct = val / total * 100
        print(f"  {chain:<14}  ${val:>10,.2f}  ({pct:.1f}%)")

    # Top 10
    print("\nTop 10 Holdings:")
    for i, (sym, data) in enumerate(sorted_agg[:10], 1):
        print(f"  {i:>2}. {sym:<10}  ${data['usd']:>10,.2f}")

    # DeFi protocols
    defi_protocols = {}
    for r in filtered:
        if r[2] == "defi":
            proto = r[3]
            defi_protocols[proto] = defi_protocols.get(proto, 0) + r[7]
    if defi_protocols:
        print("\nDeFi Protocols:")
        for proto, val in sorted(defi_protocols.items(), key=lambda x: x[1], reverse=True):
            print(f"  {proto:<20}  ${val:>10,.2f}")

    print(f"\nNote: Solana DeFi positions not included (Jupiter CAPTCHA blocked, Birdeye shows tokens only)")


if __name__ == "__main__":
    main()
